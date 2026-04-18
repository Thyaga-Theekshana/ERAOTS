"""
Report Export API — FR7: Report Generation.

Provides endpoints for exporting attendance and analytics data to:
- PDF format (using ReportLab)
- Excel format (using openpyxl)
- CSV format (native Python)

Supports:
- Attendance reports (daily, weekly, monthly)
- Employee summaries
- Department analytics
- Late arrival reports
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import joinedload
from datetime import date, datetime, timedelta
from typing import Optional, List
from io import BytesIO
import csv
import uuid

from app.core.database import get_db
from app.core.dependencies import get_current_user, require_roles
from app.core.attendance_schedule import (
    get_employee_schedule_for_date,
    compute_schedule_comparison,
)
from app.models.employee import UserAccount, Employee, Department
from app.models.attendance import AttendanceRecord

router = APIRouter(prefix="/api/reports", tags=["Reports"])


# ==================== HELPER FUNCTIONS ====================

def generate_csv(headers: List[str], rows: List[List]) -> BytesIO:
    """Generate CSV file from headers and rows."""
    output = BytesIO()
    # Write BOM for Excel compatibility
    output.write(b'\xef\xbb\xbf')
    
    # Use TextIOWrapper for CSV writer
    import io
    text_output = io.TextIOWrapper(output, encoding='utf-8', newline='')
    writer = csv.writer(text_output)
    writer.writerow(headers)
    writer.writerows(rows)
    text_output.flush()
    text_output.detach()
    
    output.seek(0)
    return output


def generate_excel(headers: List[str], rows: List[List], sheet_name: str = "Report") -> BytesIO:
    """Generate Excel file from headers and rows using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="B70100", end_color="B70100", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left")
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(str(header))
        for row in rows:
            if col_idx <= len(row):
                max_length = max(max_length, len(str(row[col_idx - 1])))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf(title: str, headers: List[str], rows: List[List], subtitle: str = "") -> BytesIO:
    """Generate PDF report using ReportLab."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#B70100")
    )
    
    # Subtitle style
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.gray
    )
    
    # Add title
    elements.append(Paragraph(title, title_style))
    if subtitle:
        elements.append(Paragraph(subtitle, subtitle_style))
    elements.append(Spacer(1, 12))
    
    # Create table
    table_data = [headers] + rows
    
    # Calculate column widths based on content
    col_widths = []
    available_width = landscape(A4)[0] - inch
    col_count = len(headers)
    
    for i in range(col_count):
        max_len = max(len(str(headers[i])), *[len(str(row[i])) if i < len(row) else 0 for row in rows])
        col_widths.append(max(min(max_len * 7, available_width / col_count), 60))
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Table style
    table_style = TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#B70100")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Body styling
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        
        # Alternating row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Footer with generation timestamp
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.gray,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(
        f"Generated by ERAOTS on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        footer_style
    ))
    
    doc.build(elements)
    output.seek(0)
    return output


# ==================== API ENDPOINTS ====================

@router.get("/attendance")
async def export_attendance_report(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    format: str = Query("excel", description="Export format: csv, excel, pdf"),
    department_id: Optional[uuid.UUID] = None,
    employee_id: Optional[uuid.UUID] = None,
    db: AsyncSession = Depends(get_db),
    current_user: UserAccount = Depends(get_current_user)
):
    """
    Export attendance report for a date range.
    
    Supports:
    - CSV format: Universal, lightweight
    - Excel format: Formatted with headers and styling
    - PDF format: Print-ready with ERAOTS branding
    """
    # Validate format
    if format not in ["csv", "excel", "pdf"]:
        raise HTTPException(status_code=400, detail="Format must be csv, excel, or pdf")
    
    # Parse dates
    try:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    # Query attendance records
    stmt = (
        select(AttendanceRecord)
        .options(joinedload(AttendanceRecord.employee).joinedload(Employee.department))
        .where(AttendanceRecord.attendance_date >= start)
        .where(AttendanceRecord.attendance_date <= end)
        .order_by(AttendanceRecord.attendance_date.desc(), AttendanceRecord.employee_id)
    )
    
    if department_id:
        stmt = stmt.join(Employee).where(Employee.department_id == department_id)
    if employee_id:
        stmt = stmt.where(AttendanceRecord.employee_id == employee_id)
    
    result = await db.execute(stmt)
    records = result.scalars().all()
    
    # Prepare data
    headers = [
        "Date", "Employee Name", "Department", "First Entry", "Last Exit",
        "Time in Building (min)", "Active Time (min)", "Break Count", "Break Duration (min)",
        "Late", "Late Duration (min)", "Overtime (min)",
        "Scheduled Start", "Scheduled End", "Scheduled Minutes", "Variance vs Schedule (min)",
        "Status"
    ]
    
    schedule_cache: dict[tuple[str, str], dict] = {}
    rows = []
    for r in records:
        dept_name = r.employee.department.name if r.employee.department else "N/A"
        cache_key = (str(r.employee_id), r.attendance_date.isoformat())
        if cache_key not in schedule_cache:
            schedule = await get_employee_schedule_for_date(db, r.employee_id, r.attendance_date)
            schedule_cache[cache_key] = compute_schedule_comparison(
                r.attendance_date,
                schedule,
                r.total_productive_time_min or 0,
            )
        comparison = schedule_cache[cache_key]

        rows.append([
            r.attendance_date.isoformat(),
            f"{r.employee.first_name} {r.employee.last_name}",
            dept_name,
            r.first_entry.strftime("%H:%M") if r.first_entry else "—",
            r.last_exit.strftime("%H:%M") if r.last_exit else "—",
            r.total_time_in_building_min or 0,
            r.total_active_time_min or 0,
            r.break_count or 0,
            r.total_break_duration_min or 0,
            "Yes" if r.is_late else "No",
            r.late_duration_min or 0,
            r.overtime_duration_min or 0,
            comparison["scheduled_start"] or "—",
            comparison["scheduled_end"] or "—",
            comparison["scheduled_minutes"] if comparison["scheduled_minutes"] is not None else "—",
            comparison["actual_vs_scheduled_variance_min"] if comparison["actual_vs_scheduled_variance_min"] is not None else "—",
            r.status or "—"
        ])
    
    # Generate file based on format
    filename = f"attendance_report_{start_date}_to_{end_date}"
    subtitle = f"Period: {start_date} to {end_date} | Records: {len(rows)}"
    
    if format == "csv":
        output = generate_csv(headers, rows)
        media_type = "text/csv"
        filename += ".csv"
    elif format == "excel":
        output = generate_excel(headers, rows, "Attendance Report")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename += ".xlsx"
    else:  # pdf
        output = generate_pdf("ERAOTS Attendance Report", headers, rows, subtitle)
        media_type = "application/pdf"
        filename += ".pdf"
    
    return StreamingResponse(
        output,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/employees")
async def export_employee_directory(
    format: str = Query("excel", description="Export format: csv, excel, pdf"),
    department_id: Optional[uuid.UUID] = None,
    status: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: UserAccount = Depends(get_current_user)
):
    """
    Export employee directory.
    
    Includes employee details with department and status information.
    """
    if format not in ["csv", "excel", "pdf"]:
        raise HTTPException(status_code=400, detail="Format must be csv, excel, or pdf")
    
    stmt = (
        select(Employee)
        .options(joinedload(Employee.department))
        .order_by(Employee.first_name, Employee.last_name)
    )
    
    if department_id:
        stmt = stmt.where(Employee.department_id == department_id)
    if status:
        stmt = stmt.where(Employee.status == status.upper())
    
    result = await db.execute(stmt)
    employees = result.scalars().all()
    
    headers = ["Employee ID", "First Name", "Last Name", "Email", "Phone", "Department", "Status", "Hire Date"]
    
    rows = []
    for e in employees:
        rows.append([
            str(e.employee_id)[:8],  # Shortened UUID
            e.first_name,
            e.last_name,
            e.email,
            e.phone or "—",
            e.department.name if e.department else "N/A",
            e.status,
            e.hire_date.isoformat() if e.hire_date else "—"
        ])
    
    filename = f"employee_directory_{date.today().isoformat()}"
    subtitle = f"Total Employees: {len(rows)} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    if format == "csv":
        output = generate_csv(headers, rows)
        media_type = "text/csv"
        filename += ".csv"
    elif format == "excel":
        output = generate_excel(headers, rows, "Employee Directory")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename += ".xlsx"
    else:
        output = generate_pdf("ERAOTS Employee Directory", headers, rows, subtitle)
        media_type = "application/pdf"
        filename += ".pdf"
    
    return StreamingResponse(
        output,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/late-arrivals")
async def export_late_arrivals_report(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    format: str = Query("excel", description="Export format: csv, excel, pdf"),
    db: AsyncSession = Depends(get_db),
    current_user: UserAccount = Depends(get_current_user)
):
    """
    Export late arrivals report.
    
    Shows employees who arrived late within the specified date range.
    """
    if format not in ["csv", "excel", "pdf"]:
        raise HTTPException(status_code=400, detail="Format must be csv, excel, or pdf")
    
    try:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    stmt = (
        select(AttendanceRecord)
        .options(joinedload(AttendanceRecord.employee).joinedload(Employee.department))
        .where(AttendanceRecord.attendance_date >= start)
        .where(AttendanceRecord.attendance_date <= end)
        .where(AttendanceRecord.is_late == True)
        .order_by(AttendanceRecord.late_duration_min.desc())
    )
    
    result = await db.execute(stmt)
    records = result.scalars().all()
    
    headers = ["Date", "Employee Name", "Department", "Arrival Time", "Late By (min)", "Status"]
    
    rows = []
    for r in records:
        dept_name = r.employee.department.name if r.employee.department else "N/A"
        rows.append([
            r.attendance_date.isoformat(),
            f"{r.employee.first_name} {r.employee.last_name}",
            dept_name,
            r.first_entry.strftime("%H:%M") if r.first_entry else "—",
            r.late_duration_min or 0,
            r.status or "—"
        ])
    
    filename = f"late_arrivals_report_{start_date}_to_{end_date}"
    subtitle = f"Period: {start_date} to {end_date} | Late Arrivals: {len(rows)}"
    
    if format == "csv":
        output = generate_csv(headers, rows)
        media_type = "text/csv"
        filename += ".csv"
    elif format == "excel":
        output = generate_excel(headers, rows, "Late Arrivals")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename += ".xlsx"
    else:
        output = generate_pdf("ERAOTS Late Arrivals Report", headers, rows, subtitle)
        media_type = "application/pdf"
        filename += ".pdf"
    
    return StreamingResponse(
        output,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/department-summary")
async def export_department_summary(
    start_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    format: str = Query("excel", description="Export format: csv, excel, pdf"),
    db: AsyncSession = Depends(get_db),
    current_user: UserAccount = Depends(get_current_user)
):
    """
    Export department attendance summary.
    
    Aggregated statistics per department for the specified period.
    """
    if format not in ["csv", "excel", "pdf"]:
        raise HTTPException(status_code=400, detail="Format must be csv, excel, or pdf")
    
    try:
        start = date.fromisoformat(start_date)
        end = date.fromisoformat(end_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    # Get all departments with employee counts and attendance stats
    departments = await db.execute(select(Department))
    depts = departments.scalars().all()
    
    headers = [
        "Department", "Employee Count", "Total Records", "Late Arrivals", 
        "Late %", "Avg Active Time (min)", "Total Overtime (min)"
    ]
    
    rows = []
    for dept in depts:
        # Count employees
        emp_count_result = await db.execute(
            select(func.count(Employee.employee_id)).where(Employee.department_id == dept.department_id)
        )
        emp_count = emp_count_result.scalar() or 0
        
        # Attendance stats for department
        stats_result = await db.execute(
            select(
                func.count(AttendanceRecord.record_id),
                func.sum(func.cast(AttendanceRecord.is_late, type_=None)),
                func.avg(AttendanceRecord.total_active_time_min),
                func.sum(AttendanceRecord.overtime_duration_min)
            )
            .join(Employee)
            .where(Employee.department_id == dept.department_id)
            .where(AttendanceRecord.attendance_date >= start)
            .where(AttendanceRecord.attendance_date <= end)
        )
        stats = stats_result.first()
        
        total_records = stats[0] or 0
        late_count = stats[1] or 0
        avg_active = round(stats[2] or 0, 1)
        total_overtime = stats[3] or 0
        late_pct = round((late_count / total_records * 100) if total_records > 0 else 0, 1)
        
        rows.append([
            dept.name,
            emp_count,
            total_records,
            late_count,
            f"{late_pct}%",
            avg_active,
            total_overtime
        ])
    
    filename = f"department_summary_{start_date}_to_{end_date}"
    subtitle = f"Period: {start_date} to {end_date} | Departments: {len(rows)}"
    
    if format == "csv":
        output = generate_csv(headers, rows)
        media_type = "text/csv"
        filename += ".csv"
    elif format == "excel":
        output = generate_excel(headers, rows, "Department Summary")
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename += ".xlsx"
    else:
        output = generate_pdf("ERAOTS Department Summary", headers, rows, subtitle)
        media_type = "application/pdf"
        filename += ".pdf"
    
    return StreamingResponse(
        output,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
